#
# This script logs into the specified mailbox and reads only the alert emails
# The Script identifies the alert emails by sender 'Alerts_us@xyz.com' 
# 										and subject starting with '**** Pending'
#
# It then picks out certain pieces of information from the mail body
# and feeds that information into an excel sheet
#
# This code uses impa-tools. Documentation in the link below
# https://pypi.python.org/pypi/imap-tools
#

from imap_tools import MailBox
import keyring
import openpyxl
import os
import re
import pandas as pd

# Initialize
mailServer = 'server.xyz.com'
ssoUsername = 'john.doe@xyz.com'

# Retrieving the password for this user from Keyring
ssoPassword = keyring.get_password('SSO', ssoUsername)

mailbox = MailBox(mailServer)
mailbox.login(ssoUsername, ssoPassword)

os.chdir('D:/John Doe/EmailDump')

# Initiate the Excel workbook for Openpyxl
wbook = 'AlertList.xlsx'
wb = openpyxl.load_workbook(wbook)
# Set the sheet named, 'Sheet' to variable named Sheet
sheet = wb.get_sheet_by_name('Sheet1')
lastRow = sheet.max_row
lastCol = sheet.max_column

# Initiate Pandas
df = pd.DataFrame()
mailData = []

# Counter
i = 0

# LIST the mail folders under the mail INBOX
for folder in mailbox.folder.list('INBOX'):
    print(folder['flags'], folder['delim'], folder['name'])

# SET the mail folder as the mails will be in it
mailbox.folder.set('INBOX/Important')

# Get the message details
for message in mailbox.fetch("FROM Alerts_us@xyz.com"):
    i += 1

    print('Subject: ', message.subject)
    print('Sent on: ', message.date)

    # Process only if this is a valid EM ALert email
    # EM Alert email subject will be like, **** Pending Sev1 SR 3-157450809 ****
    if message.subject.startswith('**** Pending'):
        # Get the date before transforming the message variable
        alertDate = message.date

        # print('Message HTML: ', message.html)
        message = message.html

        # Starting position of the text, Severity:, in the HTML body
        # print('Find 1: ', message.find('Severity:'))
        # Starting position of the text, Account:, in the HTML body
        # print('Find 2: ', message.find('Account:'))

        # Severity is the value between Severity: and Account:
        # Consider the length of the beginning text also
        severity = message[message.find('Severity: ') + len('Severity: '):message.find('Account: ')].replace('<br>', '')

        # Get Abstract - Extract from message by checking what is in between 'Abrstract: ' and 'Owner: '
        # Replace the html tag <br>
        abstract = message[message.find('Abstract:') + len('Abstract: '):message.find('Owner: ')].replace('<br>', '')
        # Split the text by ':' as text is like, [odem3];SMNH1O::Status:SMNH1O_smnh1o.com_SOA_COMPOSER_URL
        # Take the last value of the list, i.e. 3
        abstract = abstract.split(':')[3]

        # Get SR number
        sr = message[message.find('SR:') + len('SR: '):message.find('Severity: ')]
        # Using a regular expression to extract the SR number from
        # E.g. <A href="https://abcemp.us.xyz.com/epmos/mos/sp/viewSr?srNumber=3-157450809&SRProductClass=">3-157450809</A><br>
        try:
            sr = re.match(r'<.*>(.*?)</A><br>', sr, re.M | re.I).group(1)
        except:
            pass


        # Print the values
        print('Sev: ', severity)
        print('Abstract: ', abstract)
        print('SR: ', sr)
        print('Date: ', alertDate)
        # print('Message flag: ', message.flags)
        print('---------------------------------\n')

        # Transfer the values into the excel for Openpyxl
        sheet.cell(row=lastRow + i, column=2).value = alertDate
        sheet.cell(row=lastRow + i, column=3).value = sr
        sheet.cell(row=lastRow + i, column=4).value = severity
        sheet.cell(row=lastRow + i, column=5).value = abstract

        # Pandas dataframe processing
        mailData.append([alertDate, sr, severity, abstract])


# Finally, save the excel using Openpyxl
#wb.save(wbook)

# Save the Pandas Dataframe
df = pd.DataFrame(mailData,columns=['Alert Date','SR', 'Severity', 'Target'])
# print(df)

# Save to Excel from Pandas
writer = pd.ExcelWriter('AlertEmails.xlsx')
df.to_excel(writer,'Sheet1')
writer.save()

# Save to CSV from Pandas
df.to_csv('AlertEmails.csv', index=False)
